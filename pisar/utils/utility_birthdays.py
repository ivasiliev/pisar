from datetime import date
import datetime

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from classes.personnel_storage import EXCEL_DOCUMENT_SR, EXCEL_DOCUMENT_LS
from helpers.log_helper import log
from helpers.text_helper import get_month_string, get_date_str
from utils.utility_prototype import UtilityPrototype


class UtilityBirthday(UtilityPrototype):
	def __init__(self, data_model):
		super().__init__(data_model)
		if "current_date" not in data_model:
			current_date = date.today()
		else:
			current_date = data_model["current_date"]
		self.current_date = current_date
		m = self.current_date.month
		y = self.current_date.year
		m_left = m - 1
		y_left = y
		if m_left < 1:
			m_left = 12
			y_left = y - 1
		m_right = m + 2
		y_right = y
		if m_right > 12:
			m_right = 1
			y_right = y + 1
		self.date_left = datetime.date(y_left, m_left, 1)
		self.date_right = datetime.date(y_right, m_right, 1)

		log(f"Расчёт дней рождения на период [{get_date_str(self.date_left)} - {get_date_str(self.date_right)}]")

	def get_name(self):
		return "Дни рождения"

	def get_name_for_file(self):
		return f"{self.get_name()}-{self.get_date_string(self.current_date)}.xlsx"

	def render(self):
		pers_storage = self.get_pers_storage()
		rl = None
		if self.get_row_limit() > 0:
			rl = self.get_row_limit()
		all_persons = pers_storage.get_all_persons(EXCEL_DOCUMENT_SR, rl)
		all_details = pers_storage.get_all_persons(EXCEL_DOCUMENT_LS)
		persons = []
		for pers in all_persons:
			if pers.dob is None:
				continue
			nd = datetime.date(self.current_date.year, pers.dob.month, pers.dob.day)
			if self.date_left <= nd <= self.date_right:
				diff = nd - self.current_date
				pers.age = diff.days
				# TODO create function notEmpty
				if pers.unique is not None and len(pers.unique) > 0:
					for p in all_details:
						if p.unique is not None and len(p.unique) > 0 and p.unique == pers.unique:
							pers.phone = p.phone
							break
				persons.append(pers)

		log(f"Найдено {len(persons)} человек(а)")
		persons.sort(key=lambda x: x.age, reverse=False)

		captions = [
			["ФИО", 40]
			,["День рождения", 20]
			, ["Исполняется", 15]
			, ["Телефон", 15]
			, ["Должность", 20]
			, ["Где", 30]
		]

		wb = Workbook()
		ws = wb.active
		ws.title = "Дни рождения"
		f_bold = Font(bold=True)

		column_index = 0
		for c in captions:
			column_index = column_index + 1
			cell = ws.cell(row=1, column=column_index)
			cell.font = f_bold
			cell.value = c[0]
			ws.column_dimensions[get_column_letter(column_index)].width = int(c[1])

		current_year = date.today().year

		num_row = 2
		for pers in persons:
			age = current_year - pers.dob.year
			ws.cell(row=num_row, column=1, value=pers.full_name)
			ws.cell(row=num_row, column=2, value=f"{pers.dob.day} {get_month_string(pers.dob.month)}")
			ws.cell(row=num_row, column=3, value=age)
			ws.cell(row=num_row, column=4, value=pers.phone)
			ws.cell(row=num_row, column=5, value=pers.position)
			ws.cell(row=num_row, column=6, value=f"{pers.company}, {pers.platoon}, {pers.squad}")
			num_row = num_row + 1

		self.save_workbook(wb)

		super().render()
