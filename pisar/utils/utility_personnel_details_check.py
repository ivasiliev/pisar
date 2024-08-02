from openpyxl.styles import Font
from openpyxl.workbook import Workbook

from classes.person import Person
from classes.personnel_storage import EXCEL_DOCUMENT_SR, EXCEL_DOCUMENT_LS
from helpers.log_helper import log
from utils.utility_prototype import UtilityPrototype


class UtilityPersonnelDetailsCheck(UtilityPrototype):

	def get_name(self):
		return "Сверка ШР и ЛС"

	def get_name_for_file(self):
		return f"{self.get_name()}-{self.get_date_string()}.xlsx"

	def render(self):
		pers_storage = self.get_pers_storage()
		# все люди из ЛС, но ограниченное количество людей из ШР (только актуальные)
		rl = None
		rl_rep = "нет ограничений"
		if self.get_row_limit() > 0:
			rl = self.get_row_limit()
			rl_rep = str(rl)
		log(f"Анализ ШР. Ограничение по строкам: {rl_rep}.")
		persons_sr_limited = pers_storage.get_all_persons(EXCEL_DOCUMENT_SR, rl)

		log(f"Анализ ЛС. Ограничения по строкам нет.")
		persons_ls = pers_storage.get_all_persons(EXCEL_DOCUMENT_LS)

		persons_ls_clean, persons_ls_empty_unique = self.clean_pers_list(persons_ls)
		persons_sr_clean, persons_sr_empty_unique = self.clean_pers_list(persons_sr_limited)

		# есть в ЛС, нет в ШР
		pers_group_1 = self.prepare_pers_group(persons_ls_clean, persons_sr_clean)
		# есть в ШР, нет в ЛС
		pers_group_2 = self.prepare_pers_group(persons_sr_clean, persons_ls_clean)

		wb = Workbook()
		ws = wb.active
		ws.title = "Сравнение ШР и ЛС"
		ws.column_dimensions["A"].width = 60
		ws.column_dimensions["B"].width = 30

		num_row = self.print_persons_in_wb(ws, 1, "есть в ШР, нет в ЛС", pers_group_2)
		num_row = self.print_persons_in_wb(ws, num_row + 2, "есть в ЛС, нет в ШР", pers_group_1)
		num_row = self.print_persons_in_wb(ws, num_row + 2, "ЛС. Нет Личного номера", persons_ls_empty_unique)
		self.print_persons_in_wb(ws, num_row + 2, "ШР. Нет Личного номера", persons_sr_empty_unique)

		self.save_workbook(wb)

		super().render()

	def is_pers_in_list(self, pers: Person, pers_list):
		found = False
		for p in pers_list:
			if not p.get_unique_is_not_empty():
				continue
			if pers.get_unique().casefold() == p.get_unique().casefold():
				found = True
				break

		return found

	def prepare_pers_group(self, persons1, persons2):
		pers_group = []
		for p_sr in persons1:
			if not self.is_pers_in_list(p_sr, persons2):
				pers_group.append(p_sr)

		pers_group.sort(key=lambda x: x.full_name, reverse=False)
		return pers_group

	def print_persons_in_wb(self, ws, num_row_start, title, pers_group):
		num_row = num_row_start
		f_bold_14 = Font(bold=True, size=14)
		f_bold = Font(bold=True)
		ws.cell(row=num_row, column=1, value=title)
		ws[f"A{num_row}"].font = f_bold_14
		num_row = num_row + 1
		ws[f"A{num_row}"] = "ФИО"
		ws[f"B{num_row}"] = "Личный номер"
		ws[f"A{num_row}"].font = f_bold
		ws[f"B{num_row}"].font = f_bold
		num_row = num_row + 1

		if len(pers_group) > 0:
			for pers in pers_group:
				ws.cell(row=num_row, column=1, value=pers.full_name)
				ws.cell(row=num_row, column=2, value=pers.get_unique())
				num_row = num_row + 1
		else:
			ws.cell(row=num_row, column=1, value="<никого>")

		return num_row

	def clean_pers_list(self, pers_list):
		clean_list = []
		empty_unique = []
		for pers in pers_list:
			if pers.get_unique_is_not_empty():
				clean_list.append(pers)
			else:
				empty_unique.append(pers)

		return clean_list, empty_unique
