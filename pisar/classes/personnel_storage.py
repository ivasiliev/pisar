import os

import openpyxl

from classes.column_info import ColumnInfo
from classes.document_in_report import MODEL_JSON_OBJECT, MODEL_PERSONNEL_PATH, MODEL_PERSONNEL_DETAILS_PATH
from classes.excel_doc_metadata import ExcelDocMetadata
from classes.person import Person

from classes.position import Position
from helpers.file_helper import get_file_size_info
from helpers.log_helper import log
from helpers.performance_helper import PerformanceHelper
from helpers.text_helper import not_empty
import copy

openpyxl.reader.excel.warnings.simplefilter(action='ignore')

# Штатное расписание (ШР)
EXCEL_DOCUMENT_SR = 0
# Личный состав (ЛС)
EXCEL_DOCUMENT_LS = 1
COLUMN_UNIQUE_KEY = "COLUMN_UNIQUE"
COLUMN_FULL_NAME = "COLUMN_FULL_NAME"
COLUMN_DOB = "COLUMN_DOB"
COLUMN_PHONE = "COLUMN_PHONE"
COLUMN_HEIGHT = "COLUMN_HEIGHT"
COLUMN_WEIGHT = "COLUMN_WEIGHT"
COLUMN_SIGNS = "COLUMN_SIGNS"
COLUMN_TATOO = "COLUMN_TATOO"
COLUMN_ADDITIONAL_ATTRIBUTES = "COLUMN_ADDITIONAL_ATTRIBUTES"
COLUMN_HABITS = "COLUMN_HABITS"
COLUMN_PERSONAL_PERKS = "COLUMN_PERSONAL_PERKS"
COLUMN_RANK = "COLUMN_RANK"
COLUMN_POSITION = "COLUMN_POSITION"
# passports
# DNR
COLUMN_PASS_DNR = "COLUMN_PASS_DNR"
COLUMN_PASS_DNR_ISSUED = "COLUMN_PASS_DNR_ISSUED"
# RUSSIA
COLUMN_PASS_RF_NUMBER = "COLUMN_PASS_RF_NUMBER"
COLUMN_PASS_RF_ISSUE_DATE = "COLUMN_PASS_RF_ISSUE_DATE"
# кем выдан
COLUMN_PASS_RF_ISSUE_ORG = "COLUMN_PASS_RF_ISSUE_ORG"
# подразделение
COLUMN_PASS_RF_ISSUE_UNIT = "COLUMN_PASS_RF_ISSUE_UNIT"

# FOREIGN
COLUMN_PASS_FOREIGN = "COLUMN_PASS_FOREIGN"
# UKR
COLUMN_PASS_UKR = "COLUMN_PASS_UKR"
# relatives
COLUMN_FATHER_NAME = "COLUMN_FATHER_NAME"
COLUMN_MOTHER_NAME = "COLUMN_MOTHER_NAME"
COLUMN_SIBLINGS_NAME = "COLUMN_SIBLINGS_NAME"
COLUMN_SPOUSE_NAME = "COLUMN_SPOUSE_NAME"
COLUMN_FATHER_ADDRESS = "COLUMN_FATHER_ADDRESS"
COLUMN_MOTHER_ADDRESS = "COLUMN_MOTHER_ADDRESS"
COLUMN_SIBLINGS_ADDRESS = "COLUMN_SIBLINGS_ADDRESS"
COLUMN_SPOUSE_ADDRESS = "COLUMN_SPOUSE_ADDRESS"
COLUMN_FATHER_PHONE = "COLUMN_FATHER_PHONE"
COLUMN_MOTHER_PHONE = "COLUMN_MOTHER_PHONE"
COLUMN_SIBLINGS_PHONE = "COLUMN_SIBLINGS_PHONE"
COLUMN_SPOUSE_PHONE = "COLUMN_SPOUSE_PHONE"
# военкомат
COLUMN_REGISTRATION_OFFICE = "COLUMN_REGISTRATION_OFFICE"
# личные адреса
COLUMN_HOME_ADDRESS = "COLUMN_HOME_ADDRESS"
COLUMN_HOME_ADDRESS_REAL = "COLUMN_HOME_ADDRESS_REAL"
COLUMN_CITIZENSHIP = "COLUMN_CITIZENSHIP"
#
COLUMN_PLACE_OF_BIRTH = "COLUMN_PLACE_OF_BIRTH"

# документ Должности
# Полная должность
COLUMN_FULL_POSITION_NAME = "COLUMN_FULL_POSITION_NAME"
# Сокр. Должн.
COLUMN_SHORT_POSITION_NAME = "COLUMN_SHORT_POSITION_NAME"
# Должность
COLUMN_POSITION_NAME = "COLUMN_POSITION_NAME"
# Подразделение
COLUMN_UNIT = "COLUMN_UNIT"
# Подразделение 2
COLUMN_UNIT2 = "COLUMN_UNIT2"
# Взвод
COLUMN_PLATOON = "COLUMN_PLATOON"
# Отделение
COLUMN_SQUAD = "COLUMN_SQUAD"
# Воинская должность
COLUMN_MILITARY_POSITION = "COLUMN_MILITARY_POSITION"
STATUS_KIA_MIA_DES = "STATUS_KIA_MIA_DES"
RANK_ASSIGNMENT_DATE = "RANK_ASSIGNMENT_DATE"
OATH_DATE = "OATH_DATE"
OATH_TYPE = "OATH_TYPE"
BLOOD_TYPE = "BLOOD_TYPE"
ENROLLMENT_DATE = "ENROLLMENT_DATE"
SERVE_VSU_2014 = "SERVE_VSU_2014"
INCENTIVES = "INCENTIVES"
PENALTIES = "PENALTIES"
PLANS_FOR_FUTURE = "PLANS_FOR_FUTURE"
CONTRACT_DATE_FINISH = "CONTRACT_DATE_FINISH"
BANK_ACCOUNT = "BANK_ACCOUNT"
ADDRESS_MEMO = "ADDRESS_MEMO"
SOCIAL_NETWORKS = "SOCIAL_NETWORKS"
IS_DEPUTY = "IS_DEPUTY"
FAMILY_MEMBERS_COUNT = "FAMILY_MEMBERS_COUNT"
CHILDREN_COUNT = "CHILDREN_COUNT"

LS_COLUMNS = [ColumnInfo(COLUMN_RANK, "воинское звание"), ColumnInfo(COLUMN_FULL_NAME, "ФИО"), ColumnInfo(COLUMN_DOB, "дата рождения"), ColumnInfo(COLUMN_UNIQUE_KEY, "личный номер"), ColumnInfo(COLUMN_PHONE, "номер телефона")]
# TODO proper field names
# this collection also has COLUMN_UNIQUE_KEY
SR_COLUMNS = [ColumnInfo(COLUMN_UNIQUE_KEY, "личный номер"),ColumnInfo(COLUMN_FULL_POSITION_NAME, "воинское звание"), ColumnInfo(COLUMN_SHORT_POSITION_NAME, "воинское звание"),ColumnInfo(COLUMN_POSITION_NAME, "воинское звание"),ColumnInfo(COLUMN_UNIT, "воинское звание"),ColumnInfo(COLUMN_UNIT2, "воинское звание"),ColumnInfo(COLUMN_PLATOON, "воинское звание"),ColumnInfo(COLUMN_SQUAD, "воинское звание"),ColumnInfo(COLUMN_MILITARY_POSITION, "воинское звание"),ColumnInfo(COLUMN_POSITION, "воинское звание")]

class PersonnelStorage:
    def __init__(self, data_model):
        # TODO to app_settings?
        self.personnel_excel_sheet_name = "ШДС"
        self.personnel_details_excel_sheet_name = "ЛС"
        self.positions_excel_sheet_name = "должности"
        self.data_model = data_model
        self.positions_list = []
        # TODO change settings and left just one folder
        self.personnel_list_full_path = data_model[MODEL_PERSONNEL_PATH]
        self.personnel_details_full_path = data_model[MODEL_PERSONNEL_DETAILS_PATH]
        # файл Должности в том же каталоге (hardcoded file name!)
        self.positions_full_path = os.path.join(os.path.dirname(self.personnel_details_full_path), "должности.xlsx")

        # TODO calculate it automatically
        self.MAX_COLUMNS_COUNT = 150
        self.is_valid = True

        self.excel_docs = [
            self.create_metadata_for_pers_list(self.personnel_list_full_path),
            self.create_metadata_for_pers_details(self.personnel_details_full_path)]
        # self.create_metadata_for_positions(self.positions_full_path)

        for md in self.excel_docs:
            workbook = openpyxl.load_workbook(md.full_path)
            log(f"--- Сведения о файле Excel {md.full_path} ---")
            log(f"Размер: {get_file_size_info(md.full_path)}")
            log(f"Всего листов: {len(workbook.sheetnames)}")
            log("Список листов:")
            log(workbook.sheetnames)
            log("------------------------------")
            if md.sheet_name not in workbook.sheetnames:
                log(
                    f"В Excel-документе отсутствует лист '{md.sheet_name}'.")
                self.is_valid = False
            else:
                sh = workbook[md.sheet_name]
                # analyze headers
                for row in sh.iter_rows(min_row=1, min_col=1, max_row=1, max_col=md.column_count_to_search):
                    for caption_cell in row:
                        col_str = str(caption_cell.value).casefold()
                        # firstly check exact match
                        found = False
                        for col_info in md.cols:
                            if col_str == col_info.get_name():
                                col_info.index = caption_cell.col_idx
                                found = True
                                break
                        if not found:
                            # not exact match
                            for col_info in md.cols:
                                if col_info.index > -1:
                                    continue
                                if col_str.startswith(col_info.get_name()):
                                    col_info.index = caption_cell.col_idx
                                    break
                # validation
                for col_info in md.cols:
                    if not col_info.is_found():
                        pass
                        # TODO temporary switched off!
                        # log(f"Столбец '{col_info.get_name()}' не найден!")
                        # self.is_valid = False

    # Загрузка данных военнослужащего с упором на ШР. Если там находится человек, то догружаются данные из ЛС
    def find_person_by_id(self, id_person):
        id_person_str = str(id_person)
        # ШР
        pers_list_excel_doc = self.excel_docs[0]
        # ЛС
        pers_details_excel_doc = self.excel_docs[1]

        # this id must be at the first column of the personnel list Excel file
        workbook = openpyxl.load_workbook(pers_list_excel_doc.full_path)
        sh = workbook[pers_list_excel_doc.sheet_name]

        performance = PerformanceHelper()
        performance.start()
        person = None
        iteration_count_to_find_person = 0
        count_for_report = 50
        log("Поиск военнослужащего в ШР. Пожалуйста, подождите...")
        for row in sh.iter_rows(min_row=2, min_col=1, max_row=2001, max_col=self.MAX_COLUMNS_COUNT):
            person_row = None
            for cell in row:
                iteration_count_to_find_person = iteration_count_to_find_person + 1
                if cell.value is None:
                    break
                if str(cell.value) == id_person_str:
                    person_row = row
                    break
                if iteration_count_to_find_person % count_for_report == 0:
                    log(f"Обработано {iteration_count_to_find_person} строк...")
                break
            if person_row is not None:
                person = self.create_person_from_row(pers_list_excel_doc, person_row)
                break

        log(f"Количество итераций для поиска военнослужащего: {iteration_count_to_find_person}")
        performance.stop_and_print()

        # check on mandatory fields
        if person is not None and not person.get_unique_is_not_empty():
            log(f"Не задан личный номер для {person.full_name}! Продолжение работы невозможно.")
            person = None

        # let's find info in the auxiliary Excel document and replace Report Info about him
        if person is not None:
            log(
                f"Поиск военнослужащего личный номер '{person.get_unique()}' в файле Информация о личном составе. Пожалуйста, подождите...")
            performance = PerformanceHelper()
            performance.start()
            iteration_count_to_find_person = 0
            count_for_report = 50
            col_unique = pers_details_excel_doc.get_column_index(COLUMN_UNIQUE_KEY)
            workbook = openpyxl.load_workbook(pers_details_excel_doc.full_path)
            sh = workbook[pers_details_excel_doc.sheet_name]
            person_row = None
            person_unique = person.get_unique()
            for row in sh.iter_rows(min_row=2, min_col=1, max_row=2001,
                                    max_col=pers_details_excel_doc.column_count_to_search):
                for cell in row:
                    iteration_count_to_find_person = iteration_count_to_find_person + 1
                    if cell.col_idx != col_unique:
                        continue
                    if cell.value is None:
                        break
                    if str(cell.value) == person_unique:
                        person_row = row
                        break
                    if iteration_count_to_find_person % count_for_report == 0:
                        log(f"Обработано {iteration_count_to_find_person} строк...")
                    break

                if person_row is not None:
                    dm = self.data_model[MODEL_JSON_OBJECT]
                    # TODO make it automatically and easier
                    mapping = [
                        ["nationality", "COLUMN_NATIONALITY"]
                        , ["gender", "COLUMN_GENDER"]
                        , ["education", "COLUMN_EDUCATION"]
                        , ["graduation_place", "COLUMN_GRADUATION_PLACE"]
                        , ["specialization", "COLUMN_SPECIALIZATION"]
                        , ["occupation", "COLUMN_OCCUPATION"]
                        , ["foreign_languages", "COLUMN_FOREIGN_LANGUAGES"]
                        , ["awards", "COLUMN_AWARDS"]
                        , ["government_authority", "COLUMN_GOVERNMENT_AUTHORITY"]
                        , ["foreign_countries_visited", "COLUMN_FOREIGN_COUNTRIES_VISITED"]
                        , ["service_started", "COLUMN_SERVICE_STARTED"]
                        , ["place_of_birth", COLUMN_PLACE_OF_BIRTH]
                        , ["home_address", COLUMN_HOME_ADDRESS]
                        , ["home_address_real", COLUMN_HOME_ADDRESS_REAL]
                        , ["marital_status", "COLUMN_MARITAL_STATUS"]
                        , ["criminal_status", "COLUMN_CRIMINAL_STATUS"]
                        , ["phone", COLUMN_PHONE]
                        , ["height", COLUMN_HEIGHT]
                        , ["weight", COLUMN_WEIGHT]
                        , ["signs", COLUMN_SIGNS]
                        , ["tatoo", COLUMN_TATOO]
                        , ["habits", COLUMN_HABITS]
                        , ["rank", COLUMN_RANK]
                        , ["position", COLUMN_POSITION]
                        , ["additional_attributes", COLUMN_ADDITIONAL_ATTRIBUTES]
                        , ["personal_perks", COLUMN_PERSONAL_PERKS]
                        # passports
                        , ["pass_dnr", COLUMN_PASS_DNR]
                        , ["pass_dnr_issued", COLUMN_PASS_DNR_ISSUED]

                        , ["pass_rf_number", COLUMN_PASS_RF_NUMBER]
                        , ["pass_rf_issue_date", COLUMN_PASS_RF_ISSUE_DATE]
                        , ["pass_rf_issue_org", COLUMN_PASS_RF_ISSUE_ORG]
                        , ["pass_rf_issue_unit", COLUMN_PASS_RF_ISSUE_UNIT]

                        , ["pass_foreign", COLUMN_PASS_FOREIGN]
                        , ["pass_ukr", COLUMN_PASS_UKR]
                        # relatives
                        , ["father_name", COLUMN_FATHER_NAME]
                        , ["mother_name", COLUMN_MOTHER_NAME]
                        , ["siblings_name", COLUMN_SIBLINGS_NAME]
                        , ["spouse_name", COLUMN_SPOUSE_NAME]
                        , ["father_address", COLUMN_FATHER_ADDRESS]
                        , ["mother_address", COLUMN_MOTHER_ADDRESS]
                        , ["siblings_address", COLUMN_SIBLINGS_ADDRESS]
                        , ["spouse_address", COLUMN_SPOUSE_ADDRESS]
                        , ["father_phone", COLUMN_FATHER_PHONE]
                        , ["mother_phone", COLUMN_MOTHER_PHONE]
                        , ["siblings_phone", COLUMN_SIBLINGS_PHONE]
                        , ["spouse_phone", COLUMN_SPOUSE_PHONE]
                        , ["registration_office", COLUMN_REGISTRATION_OFFICE]
                        , ["citizenship", COLUMN_CITIZENSHIP]
                        #
                        , ["status_kia_mia_des", STATUS_KIA_MIA_DES]
                        , ["rank_assignment_date", RANK_ASSIGNMENT_DATE]
                        , ["oath_date", OATH_DATE]
                        , ["oath_type", OATH_TYPE]
                        , ["blood_type", BLOOD_TYPE]
                        , ["enrollment_date", ENROLLMENT_DATE]
                        , ["serve_vsu_2014", SERVE_VSU_2014]
                        , ["incentives", INCENTIVES]
                        , ["penalties", PENALTIES]
                        , ["plans_for_future", PLANS_FOR_FUTURE]
                        , ["contract_date_finish", CONTRACT_DATE_FINISH]
                        , ["bank_account", BANK_ACCOUNT]
                        , ["address_memo", ADDRESS_MEMO]
                        , ["social_networks", SOCIAL_NETWORKS]
                        , ["is_deputy", IS_DEPUTY]
                        , ["family_members_count", FAMILY_MEMBERS_COUNT]
                        , ["children_count", CHILDREN_COUNT]
                    ]

                    for m in mapping:
                        dm[m[0]] = self.find_value_in_row_by_index(person_row,
                                                                   pers_details_excel_doc.get_column_index(m[1]))

                    parents_mapping = ["father_name", "mother_name"]
                    parents_delimiter = ","
                    for p in parents_mapping:
                        n = str(dm[p])
                        if parents_delimiter in n:
                            tokens = n.split(parents_delimiter)
                            if len(tokens) >= 2:
                                dm[p] = tokens[0]

                    # select passport by priority
                    # паспорт ДНР, паспорт РФ, паспорт Украины
                    dm["passport"] = ""
                    if not_empty(dm["pass_rf_number"]):
                        dm[
                            "passport"] = f"паспорт РФ {dm['pass_rf_number']} {dm['pass_rf_issue_date']} {dm['pass_rf_issue_org']} {dm['pass_rf_issue_unit']}"
                    else:
                        if not_empty(dm["pass_dnr"]):
                            dm["passport"] = f"паспорт ДНР {dm['pass_dnr']} {dm['pass_dnr_issued']}"
                        else:
                            if not_empty(dm["pass_dnr"]):
                                dm["passport"] = f"паспорт Украины {dm['pass_ukr']}"
                    log(f"Количество итераций для поиска военнослужащего: {iteration_count_to_find_person}")
                    performance.stop_and_print()
                    break

        return person

    def find_value_in_row_by_index(self, row, index):
        result = ""
        for cell in row:
            if cell.col_idx == index:
                result = cell.value
                break
        if result is None:
            result = ""
        return result

    # ШР
    def create_metadata_for_pers_list(self, full_path):
        cols = [
            # ColumnInfo("COLUMN_COMPANY", "рота")
            ColumnInfo(COLUMN_FULL_POSITION_NAME, "полная должность")
            , ColumnInfo(COLUMN_SHORT_POSITION_NAME, "сокр. должн.")
            , ColumnInfo(COLUMN_POSITION_NAME, "должность")
            , ColumnInfo(COLUMN_UNIT, "подразделение")
            , ColumnInfo(COLUMN_UNIT2, "подразделение 2")
            , ColumnInfo(COLUMN_PLATOON, "взвод")
            , ColumnInfo(COLUMN_SQUAD, "отделение")
            , ColumnInfo(COLUMN_MILITARY_POSITION, "воинская должность")
            , ColumnInfo(COLUMN_RANK, "воинское звание")
            , ColumnInfo(COLUMN_FULL_NAME, "фио")
            , ColumnInfo(COLUMN_DOB, "дата рождения")
            , ColumnInfo(COLUMN_UNIQUE_KEY, "личный номер")
        ]

        return ExcelDocMetadata(full_path, self.personnel_excel_sheet_name, cols, 200)

    # ЛС
    def create_metadata_for_pers_details(self, full_path):
        cols = [
            ColumnInfo(COLUMN_UNIQUE_KEY, "личный номер")
            , ColumnInfo(COLUMN_RANK, "воинское звание")
            , ColumnInfo(COLUMN_FULL_NAME, "фио")
            , ColumnInfo(COLUMN_DOB, "дата рождения")
            , ColumnInfo("COLUMN_NATIONALITY", "национальность")
            , ColumnInfo("COLUMN_GENDER", "пол")
            , ColumnInfo("COLUMN_EDUCATION", "тип образования")
            , ColumnInfo("COLUMN_GRADUATION_PLACE", "учреждение(год окончания)")
            , ColumnInfo("COLUMN_SPECIALIZATION", "профессия")
            , ColumnInfo("COLUMN_OCCUPATION", "места работы")
            , ColumnInfo("COLUMN_FOREIGN_LANGUAGES", "знание иностранных языков")
            , ColumnInfo("COLUMN_AWARDS", "награды")
            , ColumnInfo("COLUMN_FOREIGN_COUNTRIES_VISITED", "какие страны посещал/посещала")
            , ColumnInfo("COLUMN_SERVICE_STARTED", "дата подписания контракта")
            , ColumnInfo(COLUMN_PLACE_OF_BIRTH, "место рождения")
            , ColumnInfo(COLUMN_HOME_ADDRESS, "адрес прописки")
            , ColumnInfo(COLUMN_HOME_ADDRESS_REAL, "фактическое место проживания")
            , ColumnInfo("COLUMN_MARITAL_STATUS", "семейное положение")
            , ColumnInfo("COLUMN_CRIMINAL_STATUS", "наличие судимостей(погашены/нет)")
            , ColumnInfo(COLUMN_PHONE, "номер телефона")
            , ColumnInfo(COLUMN_HEIGHT, "рост")
            , ColumnInfo(COLUMN_WEIGHT, "вес")
            , ColumnInfo(COLUMN_SIGNS, "особые приметы")
            , ColumnInfo(COLUMN_TATOO, "татуировки")
            , ColumnInfo(COLUMN_ADDITIONAL_ATTRIBUTES, "описание дополнительных элементов")
            , ColumnInfo(COLUMN_PERSONAL_PERKS, "индивидуальные отличительные признаки")
            , ColumnInfo(COLUMN_HABITS, "увлечения")
            # passports
            , ColumnInfo(COLUMN_PASS_DNR, "Паспорт ДНР")
            , ColumnInfo(COLUMN_PASS_DNR_ISSUED, "Кем выдан")

            , ColumnInfo(COLUMN_PASS_RF_NUMBER, "Паспорт РФ (номер)")
            , ColumnInfo(COLUMN_PASS_RF_ISSUE_DATE, "Паспорт РФ (когда выдан)")
            , ColumnInfo(COLUMN_PASS_RF_ISSUE_ORG, "Паспорт РФ Кем выдан")
            , ColumnInfo(COLUMN_PASS_RF_ISSUE_UNIT, "Подразделение")

            , ColumnInfo(COLUMN_PASS_FOREIGN, "Загранпаспорт")
            , ColumnInfo(COLUMN_PASS_UKR, "Паспорт Украины")
            # relatives
            , ColumnInfo(COLUMN_FATHER_NAME, "фио отца, дата рождения")
            , ColumnInfo(COLUMN_MOTHER_NAME, "фио матери, дата рождения")
            , ColumnInfo(COLUMN_SIBLINGS_NAME, "фио братьев/сестер, дата рождения")
            , ColumnInfo(COLUMN_SPOUSE_NAME, "фио жены/мужа, дата рождения")
            , ColumnInfo(COLUMN_FATHER_ADDRESS, "Адрес проживания отца")
            , ColumnInfo(COLUMN_MOTHER_ADDRESS, "Адрес проживания матери")
            , ColumnInfo(COLUMN_SIBLINGS_ADDRESS, "Адрес проживания братьев/сестер")
            , ColumnInfo(COLUMN_SPOUSE_ADDRESS, "Адрес проживания жены/мужа")
            , ColumnInfo(COLUMN_FATHER_PHONE, "Номер телефона отца")
            , ColumnInfo(COLUMN_MOTHER_PHONE, "Номер телефона матери")
            , ColumnInfo(COLUMN_SIBLINGS_PHONE, "Номер телефона братьев/сестер")
            , ColumnInfo(COLUMN_SPOUSE_PHONE, "Номер телефона жены/мужа")
            #
            , ColumnInfo(COLUMN_REGISTRATION_OFFICE, "Когда и каким военкоматом призван на службу")
            , ColumnInfo(COLUMN_CITIZENSHIP, "Гражданство")
            #
            , ColumnInfo(STATUS_KIA_MIA_DES, "Статус (200/БП/СОЧ)")
            , ColumnInfo(RANK_ASSIGNMENT_DATE, "Дата присвоения звания")
            , ColumnInfo(OATH_DATE, "Принятие присяги")
            , ColumnInfo(OATH_TYPE, "Контр/Мобилиз")
            , ColumnInfo(BLOOD_TYPE, "Группа крови")
            , ColumnInfo(ENROLLMENT_DATE, "Дата зачисления")
            , ColumnInfo(SERVE_VSU_2014, "Служба в ВСУ до 2014")
            , ColumnInfo(INCENTIVES, "Поощрения")
            , ColumnInfo(PENALTIES, "Взыскания")
            , ColumnInfo(PLANS_FOR_FUTURE, "Планируемый род занятий после войны")
            , ColumnInfo(CONTRACT_DATE_FINISH, "Дата окончания контракта")
            , ColumnInfo(BANK_ACCOUNT, "Банковский счет (номер банковской карты)")
            , ColumnInfo(ADDRESS_MEMO, "Адресная справка")
            , ColumnInfo(SOCIAL_NETWORKS, "Социальные сети")
            , ColumnInfo(IS_DEPUTY, "Является ли депутатом")
            , ColumnInfo(FAMILY_MEMBERS_COUNT, "Кол-во членов семьи")
            , ColumnInfo(CHILDREN_COUNT, "Количество детей")
        ]

        return ExcelDocMetadata(full_path, self.personnel_details_excel_sheet_name, cols, 200)

    # файл Должности
    def create_metadata_for_positions(self, full_path):
        cols = [
            ColumnInfo(COLUMN_FULL_POSITION_NAME, "Полная должность")
            , ColumnInfo(COLUMN_SHORT_POSITION_NAME, "Сокр. Должн.")
            , ColumnInfo(COLUMN_POSITION_NAME, "Должность")
            , ColumnInfo(COLUMN_UNIT, "Подразделение")
            , ColumnInfo(COLUMN_UNIT2, "Подразделение 2")
            , ColumnInfo(COLUMN_PLATOON, "Взвод")
            , ColumnInfo(COLUMN_SQUAD, "Отделение")
            , ColumnInfo(COLUMN_MILITARY_POSITION, "Воинская должность")
        ]

        return ExcelDocMetadata(full_path, self.positions_excel_sheet_name, cols, 200)

    # what_file=0 (ШР),what_file=1 (ЛС)
    def get_all_persons(self, what_file, row_limit=None):
        all_rows = self.read_excel_file(what_file, row_limit)
        all_persons = []
        excel_doc = self.excel_docs[what_file]
        for row in all_rows:
            all_persons.append(self.create_person_from_row(excel_doc, row))

        return all_persons

    # создает военнослужащего по ШР (только эти свойства)
    def create_person_from_row(self, excel_doc, person_row):
        # col_company = excel_doc.get_column_index("COLUMN_COMPANY")
        col_full_position_name = excel_doc.get_column_index(COLUMN_FULL_POSITION_NAME)
        col_short_position_name = excel_doc.get_column_index(COLUMN_SHORT_POSITION_NAME)
        col_position_name = excel_doc.get_column_index(COLUMN_POSITION_NAME)
        col_unit = excel_doc.get_column_index(COLUMN_UNIT)
        col_unit2 = excel_doc.get_column_index(COLUMN_UNIT2)
        col_platoon = excel_doc.get_column_index(COLUMN_PLATOON)
        col_squad = excel_doc.get_column_index(COLUMN_SQUAD)
        col_military_position = excel_doc.get_column_index(COLUMN_MILITARY_POSITION)
        col_position = excel_doc.get_column_index(COLUMN_POSITION)
        col_rank = excel_doc.get_column_index(COLUMN_RANK)
        col_full_name = excel_doc.get_column_index(COLUMN_FULL_NAME)
        col_dob = excel_doc.get_column_index(COLUMN_DOB)
        col_unique = excel_doc.get_column_index(COLUMN_UNIQUE_KEY)
        col_phone = excel_doc.get_column_index(COLUMN_PHONE)

        person = Person()
        # TODO this is incorrect because ID_SR and ID_LS are different values
        person.id_sr = self.find_value_in_row_by_index(person_row, 1)
        # TODO there is no such column anymore
        # person.company = self.find_value_in_row_by_index(person_row, col_company)

        person.full_position_name = self.find_value_in_row_by_index(person_row, col_full_position_name)
        person.short_position_name = self.find_value_in_row_by_index(person_row, col_short_position_name)
        person.position_name = self.find_value_in_row_by_index(person_row, col_position_name)
        person.unit = self.find_value_in_row_by_index(person_row, col_unit)
        person.unit2 = self.find_value_in_row_by_index(person_row, col_unit2)
        person.platoon = self.find_value_in_row_by_index(person_row, col_platoon)
        person.squad = self.find_value_in_row_by_index(person_row, col_squad)
        person.military_position = self.find_value_in_row_by_index(person_row, col_military_position)

        person.position = self.find_value_in_row_by_index(person_row, col_position)
        person.rank = self.find_value_in_row_by_index(person_row, col_rank)
        person.full_name = self.find_value_in_row_by_index(person_row, col_full_name)
        person.set_dob(self.find_value_in_row_by_index(person_row, col_dob))
        person.unique = str(self.find_value_in_row_by_index(person_row, col_unique))
        person.phone = self.find_value_in_row_by_index(person_row, col_phone)

        # normalization of a soldier name
        if person.full_name is not None:
            person.full_name = person.full_name.title()

        if person.position is not None:
            person.position = person.position.lower()
        else:
            person.position = ""
        if person.rank is not None:
            person.rank = person.rank.lower()
        else:
            person.rank = ""

        return person

    # what_file=0 (ШР),what_file=1 (ЛС)
    def read_excel_file(self, what_file, row_limit=None):
        excel_doc = self.excel_docs[what_file]
        # this id must be at the first column of the personnel list Excel file
        workbook = openpyxl.load_workbook(excel_doc.full_path)
        sh = workbook[excel_doc.sheet_name]

        # TODO
        # if None in indexes:
        #	print(f"Не удалось определить индексы столбцов! Выполнение программы прервано.")
        #	return

        performance = PerformanceHelper()
        performance.start()
        iteration_count = 0
        count_for_report = 50
        log("Просмотр списка военнослужащих. Пожалуйста, подождите...")
        max_row_value = 2001
        if row_limit is not None:
            max_row_value = row_limit
        all_rows = []
        for row in sh.iter_rows(min_row=2, min_col=1, max_row=max_row_value, max_col=self.MAX_COLUMNS_COUNT):
            if row[0].value is None:
                log(f"На строке {iteration_count + 1} в первом столбце обнаружено пустое значение. Достигнут конец таблицы.")
                break
            all_rows.append(row)

            iteration_count = iteration_count + 1
            if iteration_count % count_for_report == 0:
                log(f"Обработано {iteration_count} строк...")

        log(f"Количество итераций для просмотра списка: {iteration_count}")
        log(f"Обнаружено военнослужащих: {len(all_rows)}")
        performance.stop_and_print()

        return all_rows

    def read_excel_header(self, what_file):
        excel_doc = self.excel_docs[what_file]
        # this id must be at the first column of the personnel list Excel file
        workbook = openpyxl.load_workbook(excel_doc.full_path)
        sh = workbook[excel_doc.sheet_name]
        row_header = None
        for row in sh.iter_rows(min_row=1, min_col=1, max_row=1, max_col=self.MAX_COLUMNS_COUNT):
            row_header = row
            break

        return row_header

    # чтение всех должностей из файла Должности
    def read_positions(self):
        positions_excel_doc = self.excel_docs[2]

        # column indexes
        c1 = positions_excel_doc.get_column_index(COLUMN_FULL_POSITION_NAME)
        c2 = positions_excel_doc.get_column_index(COLUMN_SHORT_POSITION_NAME)
        c3 = positions_excel_doc.get_column_index(COLUMN_POSITION_NAME)
        c4 = positions_excel_doc.get_column_index(COLUMN_UNIT)
        c5 = positions_excel_doc.get_column_index(COLUMN_UNIT2)
        c6 = positions_excel_doc.get_column_index(COLUMN_PLATOON)
        c7 = positions_excel_doc.get_column_index(COLUMN_SQUAD)
        c8 = positions_excel_doc.get_column_index(COLUMN_MILITARY_POSITION)

        workbook = openpyxl.load_workbook(positions_excel_doc.full_path)
        sh = workbook[positions_excel_doc.sheet_name]
        log("Чтение всех должностей. Это может занять время...")
        end_of_data = False
        for row in sh.iter_rows(min_row=2, min_col=1, max_row=2001, max_col=self.MAX_COLUMNS_COUNT):
            pos = Position()
            # all cells in each row
            for cell in row:
                if cell.col_idx == 1:
                    if cell.value is None:
                        end_of_data = True
                        break
                    else:
                        pos.id = str(cell.value)
                if cell.col_idx == c1:
                    pos.full_position_name = str(cell.value)
                if cell.col_idx == c2:
                    pos.short_position_name = str(cell.value)
                if cell.col_idx == c3:
                    pos.position_name = str(cell.value)
                if cell.col_idx == c4:
                    pos.unit = str(cell.value)
                if cell.col_idx == c5:
                    pos.unit2 = str(cell.value)
                if cell.col_idx == c6:
                    pos.platoon = str(cell.value)
                if cell.col_idx == c7:
                    pos.squad = str(cell.value)
                if cell.col_idx == c8:
                    pos.military_position = str(cell.value)
            if end_of_data:
                break
            self.positions_list.append(pos)
        log(f"Чтение должностей окончено. Количество = {len(self.positions_list)}")

    # нужно создавать модель данных для каждого солдата по отдельности
    def get_data_model_by_ls_row(self, person_row, ls_excel_doc):
        dm = copy.deepcopy(self.data_model[MODEL_JSON_OBJECT])
        # TODO make it automatically and easier
        mapping = [
            ["nationality", "COLUMN_NATIONALITY"]
            , ["gender", "COLUMN_GENDER"]
            , ["education", "COLUMN_EDUCATION"]
            , ["graduation_place", "COLUMN_GRADUATION_PLACE"]
            , ["specialization", "COLUMN_SPECIALIZATION"]
            , ["occupation", "COLUMN_OCCUPATION"]
            , ["foreign_languages", "COLUMN_FOREIGN_LANGUAGES"]
            , ["awards", "COLUMN_AWARDS"]
            , ["government_authority", "COLUMN_GOVERNMENT_AUTHORITY"]
            , ["foreign_countries_visited", "COLUMN_FOREIGN_COUNTRIES_VISITED"]
            , ["service_started", "COLUMN_SERVICE_STARTED"]
            , ["place_of_birth", COLUMN_PLACE_OF_BIRTH]
            , ["home_address", COLUMN_HOME_ADDRESS]
            , ["home_address_real", COLUMN_HOME_ADDRESS_REAL]
            , ["marital_status", "COLUMN_MARITAL_STATUS"]
            , ["criminal_status", "COLUMN_CRIMINAL_STATUS"]
            , ["phone", COLUMN_PHONE]
            , ["height", COLUMN_HEIGHT]
            , ["weight", COLUMN_WEIGHT]
            , ["signs", COLUMN_SIGNS]
            , ["tatoo", COLUMN_TATOO]
            , ["habits", COLUMN_HABITS]
            , ["rank", COLUMN_RANK]
            , ["position", COLUMN_POSITION]
            , ["additional_attributes", COLUMN_ADDITIONAL_ATTRIBUTES]
            , ["personal_perks", COLUMN_PERSONAL_PERKS]
            # passports
            , ["pass_dnr", COLUMN_PASS_DNR]
            , ["pass_dnr_issued", COLUMN_PASS_DNR_ISSUED]

            , ["pass_rf_number", COLUMN_PASS_RF_NUMBER]
            , ["pass_rf_issue_date", COLUMN_PASS_RF_ISSUE_DATE]
            , ["pass_rf_issue_org", COLUMN_PASS_RF_ISSUE_ORG]
            , ["pass_rf_issue_unit", COLUMN_PASS_RF_ISSUE_UNIT]

            , ["pass_foreign", COLUMN_PASS_FOREIGN]
            , ["pass_ukr", COLUMN_PASS_UKR]
            # relatives
            , ["father_name", COLUMN_FATHER_NAME]
            , ["mother_name", COLUMN_MOTHER_NAME]
            , ["siblings_name", COLUMN_SIBLINGS_NAME]
            , ["spouse_name", COLUMN_SPOUSE_NAME]
            , ["father_address", COLUMN_FATHER_ADDRESS]
            , ["mother_address", COLUMN_MOTHER_ADDRESS]
            , ["siblings_address", COLUMN_SIBLINGS_ADDRESS]
            , ["spouse_address", COLUMN_SPOUSE_ADDRESS]
            , ["father_phone", COLUMN_FATHER_PHONE]
            , ["mother_phone", COLUMN_MOTHER_PHONE]
            , ["siblings_phone", COLUMN_SIBLINGS_PHONE]
            , ["spouse_phone", COLUMN_SPOUSE_PHONE]
            , ["registration_office", COLUMN_REGISTRATION_OFFICE]
            , ["citizenship", COLUMN_CITIZENSHIP]
            #
            , ["status_kia_mia_des", STATUS_KIA_MIA_DES]
            , ["rank_assignment_date", RANK_ASSIGNMENT_DATE]
            , ["oath_date", OATH_DATE]
            , ["oath_type", OATH_TYPE]
            , ["blood_type", BLOOD_TYPE]
            , ["enrollment_date", ENROLLMENT_DATE]
            , ["serve_vsu_2014", SERVE_VSU_2014]
            , ["incentives", INCENTIVES]
            , ["penalties", PENALTIES]
            , ["plans_for_future", PLANS_FOR_FUTURE]
            , ["contract_date_finish", CONTRACT_DATE_FINISH]
            , ["bank_account", BANK_ACCOUNT]
            , ["address_memo", ADDRESS_MEMO]
            , ["social_networks", SOCIAL_NETWORKS]
            , ["is_deputy", IS_DEPUTY]
            , ["family_members_count", FAMILY_MEMBERS_COUNT]
            , ["children_count", CHILDREN_COUNT]
        ]

        for m in mapping:
            dm[m[0]] = self.find_value_in_row_by_index(person_row,
                                                       ls_excel_doc.get_column_index(m[1]))

        parents_mapping = ["father_name", "mother_name"]
        parents_delimiter = ","
        for p in parents_mapping:
            n = str(dm[p])
            if parents_delimiter in n:
                tokens = n.split(parents_delimiter)
                if len(tokens) >= 2:
                    dm[p] = tokens[0]

        # select passport by priority
        # паспорт ДНР, паспорт РФ, паспорт Украины
        dm["passport"] = ""
        if not_empty(dm["pass_rf_number"]):
            dm[
                "passport"] = f"паспорт РФ {dm['pass_rf_number']} {dm['pass_rf_issue_date']} {dm['pass_rf_issue_org']} {dm['pass_rf_issue_unit']}"
        else:
            if not_empty(dm["pass_dnr"]):
                dm["passport"] = f"паспорт ДНР {dm['pass_dnr']} {dm['pass_dnr_issued']}"
            else:
                if not_empty(dm["pass_dnr"]):
                    dm["passport"] = f"паспорт Украины {dm['pass_ukr']}"

        return dm

    # Загрузка данных всех военнослужащих из ЛС + ШР. ЛС имеет первостепенное значение. Для каждого попытка загрузить ШР
    def load_all_persons_data(self):
        log("Загрузка всех военнослужащих ЛС + ШР. Операция может занять существенное время. Пожалуйста, подождите...")
        # result
        all_persons = []

        # ШР
        sr_excel_doc = self.excel_docs[0]
        sr_excel_doc.add_columns_in_dict(SR_COLUMNS)
        # ЛС
        ls_excel_doc = self.excel_docs[1]
        ls_excel_doc.add_columns_in_dict(LS_COLUMNS)

        col_unique_ls = ls_excel_doc.get_column_index(COLUMN_UNIQUE_KEY)
        col_unique_sr = sr_excel_doc.get_column_index(COLUMN_UNIQUE_KEY)
        if col_unique_ls is None:
            log("Не удалось найти столбец 'Личный номер' в ЛС. Выполнение программы прервано.")
            return None
        if col_unique_sr is None:
            log("Не удалось найти столбец 'Личный номер' в ШР. Выполнение программы прервано.")
            return None

        performance = PerformanceHelper()
        performance.start()

        row_num = 1

        log("Чтение ШР и ЛС. Это может занять существенное время...")

        all_sr_rows = self.read_excel_file(EXCEL_DOCUMENT_SR)
        all_ls_rows = self.read_excel_file(EXCEL_DOCUMENT_LS)

        log(f"Количество военнослужащих. ЛС: {len(all_ls_rows)}. ШР: {len(all_sr_rows)}.")

        for ls_row in all_ls_rows:
            # just check the first cell. If it is empty, we found the end of the table.
            for cell in ls_row:
                if cell.value is None:
                    log(f"Конец таблицы на строке: {row_num}")
                    break
                id_ls = self.find_value_in_row_by_index(ls_row, 1)
                if not isinstance(id_ls, int):
                    log(f"Первый столбец должен содержать только номера. Строка {row_num} содержит '{id_ls}' в первом столбце.")
                    break

                person = self.create_person_from_ls_row(ls_excel_doc, ls_row)
                if person is None:
                    log(f"Не удалось обработать военнослужащего на строке {id_ls} в ЛС. Продолжение работы невозможно.")
                    return None

                if not person.get_unique_is_not_empty():
                    log(f"Не задан личный номер для строки в ЛС номер {id_ls}! Продолжение работы невозможно.")
                    return None

                # here are all columns from LS
                person.data_model = self.get_data_model_by_ls_row(ls_row, ls_excel_doc)

                # check if the person exists in SR
                for sr_row in all_sr_rows:
                    # just check the first cell. If it is empty, we found the end of the table.
                    for cell_sr in sr_row:
                        if cell_sr.value is None:
                            break
                        unique_sr = self.find_value_in_row_by_index(sr_row, col_unique_sr)
                        if unique_sr is None:
                            id_sr = self.find_value_in_row_by_index(sr_row, 1)
                            log(f"Пустое значение для Личный номер в ШР. Строка {id_sr}. Продолжение работы невозможно.")
                            return None
                        if person.get_unique() == str(unique_sr):
                            # found!
                            person = self.add_values_from_sr_to_person(person, sr_row, sr_excel_doc)
                            person.exists_sr = True
                            break

                all_persons.append(person)
                # we don't need to read the row cell by cell
                break

            row_num = row_num + 1
        return all_persons

    # создает военнослужащего по ЛС (только эти свойства)
    def create_person_from_ls_row(self, ls_excel_doc, ls_row):
        col_rank = ls_excel_doc.get_column_index(COLUMN_RANK)
        col_full_name = ls_excel_doc.get_column_index(COLUMN_FULL_NAME)
        col_dob = ls_excel_doc.get_column_index(COLUMN_DOB)
        col_unique = ls_excel_doc.get_column_index(COLUMN_UNIQUE_KEY)
        col_phone = ls_excel_doc.get_column_index(COLUMN_PHONE)

        # TODO should be moved on Excel doc level
        if col_rank is None:
            log("ЛС. Не обнаружен столбец 'воинское звание'")
            return None

        if col_full_name is None:
            log("ЛС. Не обнаружен столбец 'ФИО'")
            return None

        if col_dob is None:
            log("ЛС. Не обнаружен столбец 'дата рождения'")
            return None

        if col_unique is None:
            log("ЛС. Не обнаружен столбец 'личный номер'")
            return None

        if col_phone is None:
            log("ЛС. Не обнаружен столбец 'номер телефона'")
            return None

        person = Person()
        person.id_sr = self.find_value_in_row_by_index(ls_row, 1)

        person.rank = self.find_value_in_row_by_index(ls_row, col_rank)
        person.full_name = self.find_value_in_row_by_index(ls_row, col_full_name)
        person.set_dob(self.find_value_in_row_by_index(ls_row, col_dob))
        person.unique = str(self.find_value_in_row_by_index(ls_row, col_unique))
        person.phone = self.find_value_in_row_by_index(ls_row, col_phone)

        # normalization of a soldier name
        if person.full_name is not None:
            person.full_name = person.full_name.title()

        if person.rank is not None:
            person.rank = person.rank.lower()
        else:
            person.rank = ""

        return person

    # добавляет военнослужащему свойства из ШР, если про него там есть запись (сводит по личному номеру)
    def add_values_from_sr_to_person(self, person, person_row, sr_excel_doc):
        col_full_position_name = sr_excel_doc.get_column_index(COLUMN_FULL_POSITION_NAME)
        col_short_position_name = sr_excel_doc.get_column_index(COLUMN_SHORT_POSITION_NAME)
        col_position_name = sr_excel_doc.get_column_index(COLUMN_POSITION_NAME)
        col_unit = sr_excel_doc.get_column_index(COLUMN_UNIT)
        col_unit2 = sr_excel_doc.get_column_index(COLUMN_UNIT2)
        col_platoon = sr_excel_doc.get_column_index(COLUMN_PLATOON)
        col_squad = sr_excel_doc.get_column_index(COLUMN_SQUAD)
        col_military_position = sr_excel_doc.get_column_index(COLUMN_MILITARY_POSITION)
        col_position = sr_excel_doc.get_column_index(COLUMN_POSITION)

        person.full_position_name = self.find_value_in_row_by_index(person_row, col_full_position_name)
        person.short_position_name = self.find_value_in_row_by_index(person_row, col_short_position_name)
        person.position_name = self.find_value_in_row_by_index(person_row, col_position_name)
        person.unit = self.find_value_in_row_by_index(person_row, col_unit)
        person.unit2 = self.find_value_in_row_by_index(person_row, col_unit2)
        person.platoon = self.find_value_in_row_by_index(person_row, col_platoon)
        person.squad = self.find_value_in_row_by_index(person_row, col_squad)
        person.military_position = self.find_value_in_row_by_index(person_row, col_military_position)
        person.position = self.find_value_in_row_by_index(person_row, col_position)

        if person.position is not None:
            person.position = person.position.lower()
        else:
            person.position = ""

        return person