import random
from openpyxl import Workbook

# ШР
file_name_sr = "personnel-demo-massive.xlsx"
# ЛС
file_name_ls = "personnel-details-massive.xlsx"

# фамилии Жуков и Жулебин будут присутствовать в ШР, но их не будет в ЛС. Специально, чтобы проверить бизнес-правило.
# special_surnames = ["Жуков", "Жулебин"]
special_surnames = []

surnames = ["Петров", "Крикунов", "Лещенко", "Исаев", "Кораблев", "Курбатов", "Магомедов", "Козлов", "Кошкин", "Ягудин",
            "Божко", "Сурков", "Щипачев", "Пархоменко", "Гладков", "Тюряев", "Васильев", "Монастырский", "Брежнев",
            "Шарапов", "Косенко", "Даутов", "Скрипалев", "Калинин", "Жуков", "Уткин", "Лебедев", "Конев", "Рублев",
            "Жулебин"]
names = ["Иван", "Петр", "Николай", "Илья", "Евгений", "Егор", "Валерий", "Алексей", "Александр", "Михаил", "Дмитрий",
         "Анатолий", "Владимир", "Игорь", "Олег", "Глеб"]
second_names = ["Иванович", "Петрович", "Николаевич", "Ильич", "Евгеньевич", "Егорович", "Валерьевич", "Алексеевич",
                "Александрович", "Дмитриевич", "Анатольевич", "Владимирович", "Игоревич", "Олегович"]
ranks = ["рядовой", "сержант", "младший лейтенант", "лейтенант", "старший лейтенант", "капитан"]
positions = ["стрелок", "гранатометчик", "пулеметчик", "радист"]
sheet_titles_sr = ["№ п/п", "Подразделение", "Рота", "Взвод", "Отделение", "Воинская должность",
                   "Воинское звание фактическое", "Личный номер", "ФИО", "Дата рождения"]
sheet_titles_ls = ["№ п/п", "Личный номер", "В/звание", "В/должность", "Статус (200/БП/СОЧ)", "Дата присвоения звания",
                   "Принятие присяги", "ФИО", "Дата рождения", "Позывной", "Номер телефона", "Национальность", "Пол",
                   "Рост", "Вес", "Группа крови", "Контр/Мобилиз", "Дата зачисления", "Служба в ВСУ до 2014",
                   "Поощрения", "Взыскания", "Тип образования", "Учреждение(год окончания)", "Профессия",
                   "Места работы", "В/У, Категория", "Знание иностранных языков", "Награды",
                   "Какие страны посещал/посещала", "Серия ВБ", "Номер ВБ", "Дата выдачи", "Военкомат",
                   "Каким военкоматом призван", "Дата подписания контракта", "Дата окончания контракта", "СНИЛС",
                   "Банковский счет", "Место рождения", "Адрес прописки", "Фактическое место проживания", "Паспорт ДНР",
                   "Кем выдан", "Паспорт РФ", "Кем выдан2", "Подразделение", "ИНН", "Паспорт Украины", "Загранпаспорт",
                   "Адресная справка", "Социальные сети", "Является ли депутатом", "Семейное положение",
                   "Кол-во членов семьи", "ФИО отца, дата рождения", "Адрес проживания отца", "Место работы отца",
                   "Номер телефона отца", "ФИО матери, дата рождения", "Адрес проживания матери", "Место работы матери",
                   "Номер телефона матери", "ФИО братьев/сестер, дата рождения", "Адрес проживания братьев/сестер",
                   "Место работы братьев/сестер", "Номер телефона братьев/сестер", "ФИО жены/мужа, дата рождения",
                   "Адрес проживания жены/мужа", "Место работы жены/мужа", "Номер телефона жены/мужа",
                   "Количество детей, ФИО", "Дата рождения детей", "Возраст детей", "Адрес проживания детей",
                   "Номер телефона детей", "Взаимоотношения в семье", "Наличие судимостей(погашены/нет)",
                   "Судимости родственников", "Вероисповедание", "Увлечения",
                   "Особые приметы(Цвет и длина волос/цвет глаз/особенность строения черепа/форма носа, ушей и т.д./наличие бороды/наличие родимых пятен, шрамов, родинок/протезов/ампутаций)",
                   "Татуировки(какие/где, описание каждой)",
                   "Описание дополнительных элементов(наручные часы, очки (контактные линзы), обереги, амулеты, кольца (в том числе обручальные), перстни, цепочки (с указанием типа металла), шнурки, молитвенная атрибутика (крестики, иконки, карманные молитвенники)",
                   "Индивидуальные отличительные признаки(размер одежды и обуви, какое имеет снаряжение и оружие, их номера, наличие особенностей и меток на форме и снаряжении)",
                   "Отношение к алкоголю", "Отношение к наркотикам", "Курение", "Попытки суицида",
                   "Участие в Б/Д(место и время, подразделение в составе которого принимал участие)",
                   "Планируемый род занятий после войны"]

max_persons_count = 500
max_company_count = 2
max_platoon_count = 3
max_squad_count = 4

min_year_birth = 1965
max_year_birth = 2003
all_letter = list("АБВГДЕЖЗИКЛМНОПРСТ")

def get_random(arr):
	return arr[random.randint(0, len(arr) - 1)]

# генерация Личного номера
def generate_unique():
	l1 = get_random(all_letter)
	l2 = get_random(all_letter)
	ns = str(random.randint(1, 999999))
	number = ns.zfill(6)
	return f"{l1}{l2}-{number}"

if __name__ == '__main__':
	wb_sr = Workbook()
	sheet_sr = wb_sr.worksheets[0]
	sheet_sr.title = "ШДС"
	for x in range(1, len(sheet_titles_sr) + 1):
		sheet_sr.cell(row=1, column=x, value=sheet_titles_sr[x - 1])

	current_person_id = 0
	ls_rows_to_record = []
	random.seed()
	print(f"Генерация {max_persons_count} военнослужащих...")
	while current_person_id < max_persons_count:
		current_person_id = current_person_id + 1

		company = random.randint(1, max_company_count)
		platoon = random.randint(1, max_platoon_count)
		squad = random.randint(1, max_squad_count)

		rank = get_random(ranks)
		position = get_random(positions)

		surname = get_random(surnames)
		name = get_random(names)
		second_name = get_random(second_names)
		full_name = f"{surname} {name} {second_name}"

		unique = generate_unique()

		day = random.randint(1, 25)
		month = random.randint(1, 12)
		year = random.randint(min_year_birth, max_year_birth)
		date_birth = f"{day}.{month}.{year}"

		# ШР
		values_to_add = [str(current_person_id), "", str(company), str(platoon), str(squad), position, rank, unique, full_name,
		                 date_birth]
		for col_ind in range(0, len(values_to_add)):
			sheet_sr.cell(row=current_person_id + 1, column=col_ind + 1, value=values_to_add[col_ind])

		# ЛС
		if surname not in special_surnames:
			ls_rows_to_record.append(
				[0, unique, rank, position, "", "", "", full_name, date_birth]
			)

		if current_person_id % 50 == 0:
			print(f" Обработано {current_person_id} записей...")
	file_to_save = f"c:\\pisar_data\\{file_name_sr}"
	wb_sr.save(file_to_save)
	print(f"Файл сохранен: {file_to_save}")

	wb_ls = Workbook()
	sheet_ls = wb_ls.worksheets[0]
	sheet_ls.title = "ЛС"
	for x in range(1, len(sheet_titles_ls) + 1):
		sheet_ls.cell(row=1, column=x, value=sheet_titles_ls[x - 1])
	row_id = 1
	for values_to_add in reversed(ls_rows_to_record):
		values_to_add[0] = row_id
		row_id = row_id + 1
		for col_ind in range(0, len(values_to_add)):
			sheet_ls.cell(row=row_id, column=col_ind + 1, value=values_to_add[col_ind])
	file_to_save = f"c:\\pisar_data\\{file_name_ls}"
	wb_ls.save(file_to_save)
	print(f"Файл сохранен: {file_to_save}")



